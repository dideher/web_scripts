import os
from openpyxl import load_workbook, Workbook #, utils

class UnexpectedCurriculumFile(Exception):
    pass

def print_curriculum_item_status(row):
    # Column numbers of interest.
    page_number_colnum = 10
    paragraph_number_colnum = 11
    justification_colnum = 12
    notes_colnum = 13

    # Read values
    page = row[page_number_colnum]
    paragraph = row[paragraph_number_colnum]
    justification = row[justification_colnum]
    if justification == "Η ομάδα προσανατολισμού δε διδάσκεται":
        justification = "H Ο.Π. δεν διδάσκεται"
    notes = row[notes_colnum]
    if notes == "ΔΕ ΔΙΔΑΣΚΕΤΑΙ":
        notes = None
    curriculum_item_list = [page, paragraph, justification, notes]
    curriculum_item_status = str()
    for entry in curriculum_item_list:
        if entry != None:
            if curriculum_item_status:
                curriculum_item_status = curriculum_item_status + "\n"+ str(entry)
            else:
                curriculum_item_status = str(entry) 
    return(curriculum_item_status)

def generate_curriculum_report(uploaded_file):
    # filename, file_extension = os.path.splitext(uploaded_file.name)
    wb = load_workbook(filename = uploaded_file)
    ws = wb.active
    if ws.title != "Sheet":
        raise UnexpectedCurriculumFile("Παρακαλώ εισάγεται το αρχείο παρακολούθησης ύλης που παρέχει το MySchool.")
        
    # wb_out = Workbook()
    wb_out = load_workbook(filename = './templates/webscripts/curriculum_report_template.xlsx')
    ws_out = wb_out.active

    number_of_entries_per_school = 25
    school_name_colnum = 7

    school_print_order = (
        '1ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '2ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟ',
        '3ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '4ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '5ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '6ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '7ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '8ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '10ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '11ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        '13ο ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        'ΠΡΟΤΥΠΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΓΙΑΣ ΒΑΡΒΑΡΑΣ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΓΙΟΣ ΜΥΡΩΝΑΣ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΓΙΩΝ ΔΕΚΑ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΡΚΑΛΟΧΩΡΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΡΧΑΝΩΝ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΑΣΗΜΙΟΥ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΒΙΑΝΝΟΥ ΗΡΑΚΛΕΙΟΥ - ΕΝΙΑΙΟ ΛΥΚΕΙΟ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΓΑΖΙΟΥ ΗΡΑΚΛΕΙΟ ΚΡΗΤΗΣ - ΔΟΜΗΝΙΚΟΣ ΘΕΟΤΟΚΟΠΟΥΛΟΣ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΓΟΥΒΩΝ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΕΠΙΣΚΟΠΗΣ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΚΑΣΤΕΛΛΙΟΥ ΠΕΔΙΑΔΑΣ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΚΡΟΥΣΩΝΑ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΛΙΜΕΝΑ ΧΕΡΣΟΝΗΣΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΑΛΙΩΝ',
        #'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΕΛΕΣΩΝ ΗΡΑΚΛΕΙΟΥ - ΜΑΛΛΙΩΤΕΙΟ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΕΛΕΣΩΝ ΗΡΑΚΛΕΙΟΥ - ΜΑΛΙΩΤΕΙΟ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΟΙΡΩΝ ΗΡΑΚΛΕΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΜΟΧΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΝΕΑΣ ΑΛΙΚΑΡΝΑΣΣΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΠΟΜΠΙΑΣ ΔΗΜΟΥ ΦΑΙΣΤΟΥ - ΓΕΛ ΠΟΜΠΙΑΣ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΤΥΜΠΑΚΙΟΥ',
        'ΗΜΕΡΗΣΙΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΧΑΡΑΚΑ ΗΡΑΚΛΕΙΟΥ',
        # 'ΚΑΛΛΙΤΕΧΝΙΚΟ ΓΥΜΝΑΣΙΟ ΗΡΑΚΛΕΙΟΥ ΜΕ ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ', 
        'ΚΑΛΛΙΤΕΧΝΙΚΟ ΣΧΟΛΕΙΟ ΗΡΑΚΛΕΙΟΥ (ΓΥΜΝΑΣΙΟ ΚΑΙ ΓΕΝΙΚΟ ΛΥΚΕΙΟ)',
        # 'Μουσικό Σχολείο Ηρακλείου\n\nΓυμνάσιο με Λυκειακές Τάξεις',
        'ΜΟΥΣΙΚΟ ΣΧΟΛΕΙΟ ΗΡΑΚΛΕΙΟΥ (ΓΥΜΝΑΣΙΟ ΚΑΙ ΓΕΝΙΚΟ ΛΥΚΕΙΟ)',
        'ΙΔΙΩΤΙΚΟ ΛΥΚΕΙΟ - ΕΚΠΑΙΔΕΥΤΗΡΙΟ "ΤΟ ΠΑΓΚΡΗΤΙΟΝ"',
        'ΕΣΠΕΡΙΝΟ ΓΕΝΙΚΟ ΛΥΚΕΙΟ ΗΡΑΚΛΕΙΟΥ',
        'ΕΣΠΕΡΙΝΟ ΓΥΜΝΑΣΙΟ ΤΥΜΠΑΚΙΟΥ ΜΕ Α , Β , Γ  ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ',
        'ΓΥΜΝΑΣΙΟ ΕΥΡΩΠΑΪΚΗΣ ΠΑΙΔΕΙΑΣ ΗΡΑΚΛΕΙΟΥ',
    )

    
    output = []
    number_of_columns = 23
    output_row = [None]*number_of_columns
    curriculum_items_titles = [None]*23

    # for title_row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    #     print(title_row)

    
    # ws_out.append(curriculum_items_titles)

    school_name = None
    school_index = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row+1, values_only=True):
        if school_name != row[7]:
            output.append(output_row)
            output_row = [None]*number_of_columns
            school_index += 1
            school_name = row[7]
            output_row[0] = school_index
            output_row[1] = school_name
            # print(f'school_name: {school_name}')
        
        if row[8] == "Γενικής Παιδείας Λυκείου" and row[9] == "Νεοελληνική Γλώσσα και Λογοτεχνία - Σύνολο Κειμένων που διδάχθηκαν":
            output_row[2] = print_curriculum_item_status(row)
        if row[8] == "Γενικής Παιδείας Λυκείου" and row[9] == "Νεοελληνική Γλώσσα και Λογοτεχνία - Σύνολο Λογ. Κειμένων που διδάχθηκαν":
            output_row[3] = print_curriculum_item_status(row)
        if row[8] == "Γενικής Παιδείας Λυκείου" and row[9] == "Νεοελληνική Γλώσσα και Λογοτεχνία - Νεοελληνική Γλώσσα – Σύνολο παραγωγών γραπτού λόγου":
            output_row[4] = print_curriculum_item_status(row)
        if row[8] == "Ανθρωπιστικές Σπουδές" and row[9] == "Αρχαία Ελληνική Γλώσσα - ΦΙΛΟΣΟΦΙΚΟΣ ΛΟΓΟΣ ΕΙΣΑΓΩΓΗ":
            output_row[5] = print_curriculum_item_status(row)
        if row[8] == "Ανθρωπιστικές Σπουδές" and row[9] == "Αρχαία Ελληνική Γλώσσα - ΦΑΚΕΛΟΣ ΥΛΙΚΟΥ":
            output_row[6] = print_curriculum_item_status(row)        
        if row[8] == "Ανθρωπιστικές Σπουδές" and row[9] == "Ιστορία":
            output_row[7] = print_curriculum_item_status(row)        
        if row[8] == "Ανθρωπιστικές Σπουδές" and row[9] == "Λατινικά - Τεύχος Α' Εισαγωγή - Μέχρι και σελ.":
            output_row[8] = print_curriculum_item_status(row)        
        if row[8] == "Ανθρωπιστικές Σπουδές" and row[9] == "Λατινικά - Τεύχος Α' Aριθμός τελ. μαθήματος":
            output_row[9] = print_curriculum_item_status(row)        
        if row[8] == "Ανθρωπιστικές Σπουδές" and row[9] == "Λατινικά - Τεύχος Β' Aριθμός τελ. μαθήματος":
            output_row[10] = print_curriculum_item_status(row)        
        if row[8] == "Θετικές Σπουδές και Υγείας - Θετικές και Τεχνολογικές Επιστήμες" and row[9] == "Μαθηματικά":
            output_row[11] = print_curriculum_item_status(row)
        output_row[12] = " "
        if row[8] == "Θετικές Σπουδές και Υγείας - Επιστήμες Υγείας και Ζωής" and row[9] == "Φυσική Τεύχος Β'":
            output_row[13] = print_curriculum_item_status(row)
        if row[8] == "Θετικές Σπουδές και Υγείας - Επιστήμες Υγείας και Ζωής" and row[9] == "Φυσική Τεύχος Γ'":
            output_row[14] = print_curriculum_item_status(row)
        if row[8] == "Θετικές Σπουδές και Υγείας - Επιστήμες Υγείας και Ζωής" and row[9] == "Χημεία Τεύχος Α'":
            output_row[15] = print_curriculum_item_status(row)
        if row[8] == "Θετικές Σπουδές και Υγείας - Επιστήμες Υγείας και Ζωής" and row[9] == "Χημεία Τεύχος Β'":
            output_row[16] = print_curriculum_item_status(row)    
        if row[8] == "Θετικές Σπουδές και Υγείας - Επιστήμες Υγείας και Ζωής" and row[9] == "Βιολογία Τεύχος Α'":
            output_row[17] = print_curriculum_item_status(row)
        if row[8] == "Θετικές Σπουδές και Υγείας - Επιστήμες Υγείας και Ζωής" and row[9] == "Βιολογία Τεύχος Β'":
            output_row[18] = print_curriculum_item_status(row)
        if row[8] == "Σπουδές Οικονομίας και Πληροφορικής" and row[9] == "Μαθηματικά":
            output_row[19] = print_curriculum_item_status(row)
        if row[8] == "Σπουδές Οικονομίας και Πληροφορικής" and row[9] == "Πληροφορική / Α.Ε.Π.Π. Βιβλίο Μαθητή":
            output_row[20] = print_curriculum_item_status(row)
        if row[8] == "Σπουδές Οικονομίας και Πληροφορικής" and row[9] == "Πληροφορική / Πληρ. Συμπλ. Εκπ. Υλ.":
            output_row[21] = print_curriculum_item_status(row)
        if row[8] == "Σπουδές Οικονομίας και Πληροφορικής" and row[9] == "Οικονομία":
            output_row[22] = print_curriculum_item_status(row)

    new_index = 1
    for e in school_print_order:
        r = [r for r in output if r[1] == e]
        if r:
            if r[0][0]:
                r[0][0] = new_index
            new_index += 1
            ws_out.append(r[0])
    wb_out.save('/tmp/curriculum_report.xlsx')